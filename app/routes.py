import logging
from flask import render_template, redirect, url_for, flash, request, jsonify, session
from flask_login import login_user, logout_user, login_required, current_user
from urllib.parse import urlparse
from app import app, db
from app.models import User, Income, Expense, Budget, Student, FeePayment
from app.forms import AdmissionForm, FeePaymentForm, StudentSearchForm, MonthlyFeeReportForm
from werkzeug.security import check_password_hash
from werkzeug.exceptions import HTTPException

# Configure logging
logger = logging.getLogger(__name__)

# Global error handler
@app.errorhandler(Exception)
def handle_error(e):
    code = 500
    if isinstance(e, HTTPException):
        code = e.code
    
    logger.error(f'Unhandled Exception: {str(e)}', exc_info=True)
    
    return jsonify(error=str(e)), code

# Catch-all route to handle method not allowed
@app.route('/', methods=['GET', 'POST'])
def catch_all():
    logger.warning(f'Catch-all route hit with method: {request.method}')
    if current_user.is_authenticated:
        total_income = sum(income.amount for income in Income.query.all())
        total_expenses = sum(expense.amount for expense in Expense.query.all())
        return render_template('dashboard.html', 
                               total_income=total_income, 
                               total_expenses=total_expenses)
    return render_template('login.html')
from werkzeug.utils import secure_filename
import os
import openpyxl
from datetime import datetime

@app.route('/')
def index():
    print('Current user:', current_user)
    print('Is authenticated:', getattr(current_user, 'is_authenticated', None))
    if current_user.is_authenticated:
        # Calculate totals for the dashboard
        total_income = sum(income.amount for income in Income.query.all())
        total_expenses = sum(expense.amount for expense in Expense.query.all())
        
        # Get recent incomes and expenses for display
        recent_incomes = Income.query.order_by(Income.date.desc()).limit(5).all()
        recent_expenses = Expense.query.order_by(Expense.date.desc()).limit(5).all()
        
        # Add student and fee stats for the dashboard
        total_students = Student.query.count()
        total_fee_collected = db.session.query(db.func.sum(FeePayment.amount)).scalar() or 0
        
        # Get recent fee payments
        recent_payments = FeePayment.query.order_by(FeePayment.payment_date.desc()).limit(5).all()
        
        from datetime import datetime
        current_year = datetime.now().year
        current_month = datetime.now().month
        
        return render_template('dashboard.html', 
                               total_income=total_income, 
                               total_expenses=total_expenses,
                               recent_incomes=recent_incomes,
                               recent_expenses=recent_expenses,
                               total_students=total_students,
                               total_fee_collected=total_fee_collected,
                               recent_payments=recent_payments,
                               current_month=current_month,
                               now=lambda: datetime.now())
    return render_template('login.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    # Don't show login page if already logged in
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    
    print(f"\n============= LOGIN ROUTE ==============")
    print(f"Request method: {request.method}")
    print(f"Current user: {current_user}")
    
    try:
        if request.method == 'POST':
            username = request.form.get('username', '')
            password = request.form.get('password', '')
            
            print(f"Login attempt - Username: {username}, Password length: {len(password)}")
            
            if not username or not password:
                flash('Username and password are required', 'danger')
                return render_template('login.html'), 400
            
            # For debugging - verify admin exists
            print("Checking if admin exists...")
            admin_user = User.query.filter_by(username='admin').first()
            if admin_user:
                print(f"Admin user exists - ID: {admin_user.id}, Email: {admin_user.email}")
                print(f"Password hash: {admin_user.password_hash[:10]}...")
            else:
                print("No admin user found in database!")
            
            # Check user
            user = User.query.filter_by(username=username).first()
            
            if not user:
                flash('Invalid username or password', 'danger')
                return render_template('login.html')
            
            # Check password
            is_password_correct = user.check_password(password)
            print(f"Password check result: {is_password_correct}")
            
            if is_password_correct:
                # Login the user with remember=True for persistence
                login_user(user, remember=True)
                print(f"User logged in successfully: {current_user.is_authenticated}")
                
                # Important: Use next_page pattern to handle redirects properly
                next_page = request.args.get('next')
                if not next_page or urlparse(next_page).netloc != '':
                    next_page = url_for('index')
                
                flash('Login successful!', 'success')
                return redirect(next_page)
            else:
                flash('Invalid username or password', 'danger')
        
        # GET request - render the login form
        return render_template('login.html')
    
    except Exception as e:
        print(f"Login error: {str(e)}")
        flash('An unexpected error occurred. Please try again.', 'danger')
        return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route('/income', methods=['GET', 'POST'])
@login_required
def manage_income():
    if request.method == 'POST':
        source = request.form['source']
        amount = float(request.form['amount'])
        description = request.form['description']
        
        new_income = Income(source=source, amount=amount, description=description)
        db.session.add(new_income)
        db.session.commit()
        flash('Income added successfully!')
        return redirect(url_for('manage_income'))
    
    incomes = Income.query.order_by(Income.date.desc()).all()
    return render_template('income.html', incomes=incomes)

@app.route('/expenses', methods=['GET', 'POST'])
@login_required
def manage_expenses():
    if request.method == 'POST':
        category = request.form['category']
        amount = float(request.form['amount'])
        description = request.form['description']
        
        new_expense = Expense(category=category, amount=amount, description=description)
        db.session.add(new_expense)
        db.session.commit()
        flash('Expense added successfully!')
        return redirect(url_for('manage_expenses'))
    
    expenses = Expense.query.order_by(Expense.date.desc()).all()
    return render_template('expenses.html', expenses=expenses)

@app.route('/budget', methods=['GET', 'POST'])
@login_required
def manage_budget():
    if request.method == 'POST':
        category = request.form['category']
        allocated_amount = float(request.form['allocated_amount'])
        fiscal_year = int(request.form['fiscal_year'])
        description = request.form['description']
        
        new_budget = Budget(category=category, 
                            allocated_amount=allocated_amount, 
                            fiscal_year=fiscal_year, 
                            description=description)
        db.session.add(new_budget)
        db.session.commit()
        flash('Budget added successfully!')
        return redirect(url_for('manage_budget'))
    
    budgets = Budget.query.order_by(Budget.fiscal_year.desc()).all()
    return render_template('budget.html', budgets=budgets)

@app.route('/reports')
@login_required
def financial_reports():
    # Implement financial reporting logic
    incomes = Income.query.all()
    expenses = Expense.query.all()
    budgets = Budget.query.all()
    
    return render_template('reports.html', 
                           incomes=incomes, 
                           expenses=expenses, 
                           budgets=budgets)

@app.route('/import', methods=['GET', 'POST'])
@login_required
def import_excel():
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('No file part')
            return redirect(request.url)
        
        file = request.files['file']
        
        if file.filename == '':
            flash('No selected file')
            return redirect(request.url)
        
        if file:
            filename = secure_filename(file.filename)
            file_path = os.path.join(app.root_path, 'uploads', filename)
            os.makedirs(os.path.dirname(file_path), exist_ok=True)
            file.save(file_path)
            
            try:
                workbook = openpyxl.load_workbook(file_path)
                
                # Import Income
                if 'Income' in workbook.sheetnames:
                    income_sheet = workbook['Income']
                    for row in income_sheet.iter_rows(min_row=2, values_only=True):
                        if row[0]:  # Check if row is not empty
                            new_income = Income(
                                source=str(row[0]),
                                amount=float(row[1]),
                                date=datetime.strptime(str(row[2]), '%Y-%m-%d') if row[2] else datetime.now(),
                                description=str(row[3]) if row[3] else ''
                            )
                            db.session.add(new_income)
                
                # Import Expenses
                if 'Expenses' in workbook.sheetnames:
                    expense_sheet = workbook['Expenses']
                    for row in expense_sheet.iter_rows(min_row=2, values_only=True):
                        if row[0]:  # Check if row is not empty
                            new_expense = Expense(
                                category=str(row[0]),
                                amount=float(row[1]),
                                date=datetime.strptime(str(row[2]), '%Y-%m-%d') if row[2] else datetime.now(),
                                description=str(row[3]) if row[3] else ''
                            )
                            db.session.add(new_expense)
                
                # Import Budget
                if 'Budget' in workbook.sheetnames:
                    budget_sheet = workbook['Budget']
                    for row in budget_sheet.iter_rows(min_row=2, values_only=True):
                        if row[0]:  # Check if row is not empty
                            new_budget = Budget(
                                category=str(row[0]),
                                allocated_amount=float(row[1]),
                                fiscal_year=int(row[2]),
                                description=str(row[3]) if row[3] else ''
                            )
                            db.session.add(new_budget)
                
                db.session.commit()
                flash('Excel file imported successfully!')
            except Exception as e:
                db.session.rollback()
                flash(f'Error importing file: {str(e)}')
            
            # Clean up uploaded file
            os.remove(file_path)
            
            return redirect(url_for('index'))
    
    return render_template('import.html')


# Student Management Routes
@app.route('/students', methods=['GET'])
@login_required
def list_students():
    search_form = StudentSearchForm()
    search_query = request.args.get('search_query', '')
    
    if search_query:
        # Search by name or GR number
        students = Student.query.filter(
            (Student.name.contains(search_query)) | 
            (Student.gr_number.contains(search_query))
        ).all()
    else:
        students = Student.query.all()
    
    return render_template('students/list.html', 
                           students=students, 
                           search_form=search_form)


@app.route('/students/add', methods=['GET', 'POST'])
@login_required
def add_student():
    form = AdmissionForm()
    
    if form.validate_on_submit():
        student = Student(
            gr_number=form.gr_number.data,
            name=form.name.data,
            father_name=form.father_name.data,
            class_name=form.class_name.data,
            nationality=form.nationality.data,
            religion=form.religion.data,
            gender=form.gender.data,
            dob=form.dob.data,
            cnic=form.cnic.data,
            address=form.address.data,
            guardian_type=form.guardian_type.data,
            guardian_name=form.guardian_name.data,
            guardian_relation=form.guardian_relation.data,
            guardian_phone=form.guardian_phone.data,
            guardian_email=form.guardian_email.data,
            guardian_occupation=form.guardian_occupation.data,
            previous_school=form.previous_school.data,
            last_exam_passed=form.last_exam_passed.data,
            last_exam_marks=form.last_exam_marks.data,
            last_exam_grade=form.last_exam_grade.data,
            last_exam_year=form.last_exam_year.data,
            transfer_certificate=form.transfer_certificate.data
        )
        
        db.session.add(student)
        db.session.commit()
        
        flash(f'Student {student.name} has been added successfully!', 'success')
        return redirect(url_for('list_students'))
    
    return render_template('students/add.html', form=form, title='New Student Admission')


@app.route('/students/<int:student_id>', methods=['GET'])
@login_required
def view_student(student_id):
    student = Student.query.get_or_404(student_id)
    fee_payments = FeePayment.query.filter_by(student_id=student_id).order_by(FeePayment.year.desc(), FeePayment.month.desc()).all()
    
    # Generate fee calendar for current year
    current_year = datetime.now().year
    fee_calendar = []
    
    for month in range(1, 13):
        status = student.get_fee_status(month, current_year)
        fee_calendar.append({
            'month': month,
            'month_name': [
                'January', 'February', 'March', 'April', 'May', 'June',
                'July', 'August', 'September', 'October', 'November', 'December'
            ][month-1],
            'status': status
        })
    
    monthly_fee = student.calculate_monthly_fee()
    
    return render_template('students/view.html', 
                           student=student, 
                           fee_payments=fee_payments,
                           fee_calendar=fee_calendar,
                           monthly_fee=monthly_fee)


@app.route('/students/<int:student_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_student(student_id):
    student = Student.query.get_or_404(student_id)
    form = AdmissionForm(obj=student)
    
    if form.validate_on_submit():
        student.gr_number = form.gr_number.data
        student.name = form.name.data
        student.father_name = form.father_name.data
        student.class_name = form.class_name.data
        student.nationality = form.nationality.data
        student.religion = form.religion.data
        student.gender = form.gender.data
        student.dob = form.dob.data
        student.cnic = form.cnic.data
        student.address = form.address.data
        student.guardian_type = form.guardian_type.data
        student.guardian_name = form.guardian_name.data
        student.guardian_relation = form.guardian_relation.data
        student.guardian_phone = form.guardian_phone.data
        student.guardian_email = form.guardian_email.data
        student.guardian_occupation = form.guardian_occupation.data
        student.previous_school = form.previous_school.data
        student.last_exam_passed = form.last_exam_passed.data
        student.last_exam_marks = form.last_exam_marks.data
        student.last_exam_grade = form.last_exam_grade.data
        student.last_exam_year = form.last_exam_year.data
        student.transfer_certificate = form.transfer_certificate.data
        
        db.session.commit()
        
        flash(f'Student {student.name} has been updated successfully!', 'success')
        return redirect(url_for('view_student', student_id=student.id))
    
    return render_template('students/edit.html', form=form, student=student, title='Edit Student')


@app.route('/students/<int:student_id>/delete', methods=['POST'])
@login_required
def delete_student(student_id):
    student = Student.query.get_or_404(student_id)
    
    # Delete all fee payments related to this student
    FeePayment.query.filter_by(student_id=student_id).delete()
    
    db.session.delete(student)
    db.session.commit()
    
    flash(f'Student {student.name} has been deleted!', 'success')
    return redirect(url_for('list_students'))


# Fee Management Routes
@app.route('/fees', methods=['GET'])
@login_required
def fee_dashboard():
    # Get some stats for the dashboard
    total_students = Student.query.count()
    fees_collected_current_month = db.session.query(db.func.sum(FeePayment.amount)).filter(
        FeePayment.month == datetime.now().month,
        FeePayment.year == datetime.now().year,
        FeePayment.payment_status == 'Paid'
    ).scalar() or 0
    
    # Get students with pending fees
    current_month = datetime.now().month
    current_year = datetime.now().year
    
    # Subquery to find students who have paid this month
    paid_students = db.session.query(FeePayment.student_id).filter(
        FeePayment.month == current_month,
        FeePayment.year == current_year,
        FeePayment.payment_status == 'Paid'
    ).subquery()
    
    # Find students who haven't paid
    pending_students = Student.query.filter(
        ~Student.id.in_(paid_students)
    ).all()
    
    return render_template('fees/dashboard.html',
                           total_students=total_students,
                           fees_collected=fees_collected_current_month,
                           pending_students=pending_students)


@app.route('/fees/collect', methods=['GET', 'POST'])
@login_required
def collect_fee():
    form = FeePaymentForm()
    
    # Populate student choices
    students = Student.query.all()
    form.student.choices = [(s.id, f'{s.name} ({s.gr_number}) - {s.class_name}') for s in students]
    
    if form.validate_on_submit():
        student = Student.query.get(form.student.data)
        
        # Check if fee for this month/year already exists
        existing_payment = FeePayment.query.filter_by(
            student_id=student.id,
            month=form.month.data,
            year=form.year.data
        ).first()
        
        if existing_payment:
            flash(f'Fee record already exists for {student.name} for {form.month.data}/{form.year.data}. Please edit the existing record.', 'warning')
            return redirect(url_for('view_student', student_id=student.id))
        
        # Create new payment
        payment = FeePayment(
            student_id=student.id,
            month=form.month.data,
            year=form.year.data,
            amount=form.amount.data,
            payment_date=form.payment_date.data,
            payment_status=form.payment_status.data,
            payment_method=form.payment_method.data,
            receipt_number=form.receipt_number.data,
            remarks=form.remarks.data
        )
        
        db.session.add(payment)
        db.session.commit()
        
        flash(f'Fee payment for {student.name} has been recorded successfully!', 'success')
        return redirect(url_for('view_student', student_id=student.id))
    
    # Pre-fill some fields with default values
    if not form.month.data:
        form.month.data = datetime.now().month
    if not form.year.data:
        form.year.data = datetime.now().year
    if not form.payment_date.data:
        form.payment_date.data = datetime.now().date()
    
    return render_template('fees/collect.html', form=form, title='Collect Fee')


@app.route('/fees/report', methods=['GET', 'POST'])
@login_required
def fee_report():
    form = MonthlyFeeReportForm()
    
    # Populate class filter choices
    class_choices = [(c.class_name, c.class_name) for c in db.session.query(Student.class_name).distinct()]
    form.class_filter.choices = [('', 'All Classes')] + class_choices
    
    report_data = None
    total_amount = 0
    total_paid = 0
    total_pending = 0
    
    if form.validate_on_submit() or request.args.get('month'):
        # Get filter parameters
        month = form.month.data if form.validate_on_submit() else int(request.args.get('month'))
        year = form.year.data if form.validate_on_submit() else int(request.args.get('year'))
        class_filter = form.class_filter.data if form.validate_on_submit() else request.args.get('class_filter')
        
        # Build query
        query = Student.query
        
        if class_filter:
            query = query.filter_by(class_name=class_filter)
        
        students = query.all()
        report_data = []
        
        for student in students:
            # Calculate fee amount
            fee_amount = student.calculate_monthly_fee()
            total_amount += fee_amount
            
            # Get payment status
            payment = FeePayment.query.filter_by(
                student_id=student.id,
                month=month,
                year=year
            ).first()
            
            status = 'Unpaid'
            paid_amount = 0
            
            if payment:
                status = payment.payment_status
                paid_amount = payment.amount if payment.payment_status == 'Paid' else 0
            
            if status == 'Paid':
                total_paid += fee_amount
            else:
                total_pending += fee_amount
            
            report_data.append({
                'student': student,
                'fee_amount': fee_amount,
                'status': status,
                'paid_amount': paid_amount
            })
    
    # Pre-fill form with current month/year
    if not form.month.data:
        form.month.data = datetime.now().month
    if not form.year.data:
        form.year.data = datetime.now().year
    
    return render_template('fees/report.html', 
                           form=form, 
                           report_data=report_data,
                           total_amount=total_amount,
                           total_paid=total_paid,
                           total_pending=total_pending,
                           title='Fee Collection Report')

